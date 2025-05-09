from customtkinter import *
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageDraw, ImageOps
import subprocess
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pandas as pd
import numpy as np


# ==================== Window Setup ====================
window = CTk()
window.title("Admin Panel - Student Management System")
height = 825
width = 1300
x = (window.winfo_screenwidth()//2)-(width//2) 
y = (window.winfo_screenheight()//2)-(height//2) 
window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
window.resizable(False, False)
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

style = ttk.Style()
style.theme_use('clam')

df_rows = []

# ==================== Functions ====================
# User Account Show Data Table
def user_account_show_data():
    global df_rows

    opening_file = filedialog.askopenfilename(title="Select File", filetypes=[("Excel Files", "*.xlsx")])
    if not opening_file:
        return

    try:
        df = pd.read_excel(opening_file)
        df_rows = df.to_numpy().tolist()
        show_all_user_account_data()
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open file: {e}")
        return

# Personal Details Show Data Table
def personal_details_show_data():
    global rset
    try:
        personal_search_entry.delete(0, "end")

        sections = personal_section_option.get().strip()
        if sections not in ["1A", "1B", "1C"]:
            messagebox.showerror("Error", "Please select a valid section.")
            return
        
        new_sheet_name = f"{sections} Enrollment Information "
        file_path_to_excel = "user_account_data.xlsx"

        wb = load_workbook(file_path_to_excel)
        if new_sheet_name not in wb.sheetnames:
            messagebox.showerror("Error", f"Sheet '{new_sheet_name}' not found in the workbook.")
            return
        
        ws = wb[new_sheet_name]
        rset = ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=26, values_only=True)
        rset = [r for r in rset]
        wb.close()

        personal_student_table.delete(*personal_student_table.get_children())
        for row in rset:
            personal_student_table.insert("", "end", values=row)

    except FileNotFoundError:
        messagebox.showerror("Error", "File not found.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open file: {e}")

# Family Details Show Data Table
def family_details_show_data():
    global rset
    try:
        family_search_entry.delete(0, "end")

        sections = family_section_option.get().strip()
        if sections not in ["1A", "1B", "1C"]:
            messagebox.showerror("Error", "Please select a valid section.")
            return
        
        new_sheet_name = f"{sections} Enrollment Information "
        file_path_to_excel = "user_account_data.xlsx"

        wb = load_workbook(file_path_to_excel)
        if new_sheet_name not in wb.sheetnames:
            messagebox.showerror("Error", f"Sheet '{new_sheet_name}' not found in the workbook.")
            return
        
        ws = wb[new_sheet_name]
        rset = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=28,max_col=42, values_only=True)
        rset = [r for r in rset]
        wb.close()

        family_student_table.delete(*family_student_table.get_children())
        for row in rset:
            family_student_table.insert("", "end", values=row)

    except FileNotFoundError:
        messagebox.showerror("Error", "File not found.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open file: {e}")

# Educational Details Show Data Table
def educational_details_show_data():
    global rset
    try:
        educ_search_entry.delete(0, "end")

        sections = educ_section_option.get().strip()
        if sections not in ["1A", "1B", "1C"]:
            messagebox.showerror("Error", "Please select a valid section.")
            return
        
        new_sheet_name = f"{sections} Enrollment Information "
        file_path_to_excel = "user_account_data.xlsx"

        wb = load_workbook(file_path_to_excel)
        if new_sheet_name not in wb.sheetnames:
            messagebox.showerror("Error", f"Sheet '{new_sheet_name}' not found in the workbook.")
            return
        
        ws = wb[new_sheet_name]
        rset = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=44, max_col=59, values_only=True)
        rset = [r for r in rset]
        wb.close()

        educ_student_table.delete(*educ_student_table.get_children())
        for row in rset:
            educ_student_table.insert("", "end", values=row)

    except FileNotFoundError:
        messagebox.showerror("Error", "File not found.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open file: {e}")

# Search For User
def user_account_search():
    search_option = user_account_search_option.get()
    search_query = user_account_search_entry.get().lower()

    if not search_query:
        messagebox.showwarning("Input Error", "Please enter a search query.")
        return
    
    filtered_data = []
    for row in df_rows:
        if search_option == "Username" and search_query in row[3].lower():
            filtered_data.append(row)
        elif search_option == "Email" and search_query in row[2].lower():
            filtered_data.append(row)
        elif search_option == "First Name" and search_query in row[0].lower():
            filtered_data.append(row)
        elif search_option == "Last Name" and search_query in row[1].lower():
            filtered_data.append(row)

    user_account_student_table.delete(*user_account_student_table.get_children())

    for row in filtered_data:
        user_account_student_table.insert("", "end", values=row)

    if not filtered_data:
        messagebox.showinfo("No Results", "No matching records found.")
        user_account_show_data()

def show_all_user_account_data():
    global df_rows

    user_account_search_entry.delete(0, "end")

    if not df_rows:
        messagebox.showwarning("Warning", "No data found. Please upload a file first.")
        return

    try:
        user_account_student_table.delete(*user_account_student_table.get_children())
        for row in df_rows:
            user_account_student_table.insert("", "end", values=row)
    except Exception as e:
        messagebox.showerror("Error", f"Unexpected error: {e}")


#Search For Student
def personal_details_search():
    search_option = personal_search_option.get()
    search_query = personal_search_entry.get().lower()

    if not search_query:
        messagebox.showwarning("Input Error", "Please enter a search query.")
        return
    
    filtered_data = []
    for row in rset:
        if search_option == "Student ID" and search_query in str(row[0]).lower():  # Column 0 is "Student ID"
            filtered_data.append(row)
        elif search_option == "Surname" and search_query in str(row[4]).lower():  # Column 4 is "Surname"
            filtered_data.append(row)

    personal_student_table.delete(*personal_student_table.get_children())

    for row in filtered_data:
        personal_student_table.insert("", "end", values=row)

    if not filtered_data:
        messagebox.showinfo("No Results", "No matching records found.")
        personal_details_show_data()

# Search For Family ================= NOT FUNCTIONING PROPERLY =================
def family_details_search():
    search_option = family_search_option.get()
    search_query = family_search_entry.get().strip().lower()

    if not search_query:
        messagebox.showwarning("Input Error", "Please enter a search query.")
        return

    filtered_data = []
    for row in rset:
        if len(row) > 37:
            print(f"Row data: {row}")
            
            if search_option == "Father's Name" and row[0] and search_query in str(row[0]).lower():
                filtered_data.append(row)
            elif search_option == "Mother's Name" and row[5] and search_query in str(row[5]).lower():
                filtered_data.append(row)
            elif search_option == "Guardian's Name" and row[10] and search_query in str(row[10]).lower():
                filtered_data.append(row)

    family_student_table.delete(*family_student_table.get_children())

    for row in filtered_data:
        family_student_table.insert("", "end", values=row)

    if not filtered_data:
        messagebox.showinfo("No Results", "No matching records found.")
        family_details_show_data()

# Search For Education
def educational_details_search():
    search_option = educ_search_option.get()
    search_query = educ_search_entry.get().lower()

    if not search_query:
        messagebox.showwarning("Input Error", "Please enter a search query.")
        return
    
    filtered_data = []
    for row in rset:
        if search_option == "ELEM" and search_query in str(row[0]).lower():
            filtered_data.append(row)
        elif search_option == "JHS" and search_query in str(row[4]).lower():
            filtered_data.append(row)
        elif search_option == "SHS" and search_query in str(row[8]).lower():
            filtered_data.append(row)
        elif search_option == "COL" and search_query in str(row[12]).lower():
            filtered_data.append(row)

    educ_student_table.delete(*educ_student_table.get_children())

    for row in filtered_data:
        educ_student_table.insert("", "end", values=row)

    if not filtered_data:
        messagebox.showinfo("No Results", "No matching records found.")
        educational_details_show_data()

# Delete Button
def delete_selected_row():
    tables = [
        user_account_student_table,
        personal_student_table,
        family_student_table,
        educ_student_table
    ]
    
    selected_found = False
    for table in tables:
        selected_item = table.selection()
        if selected_item:
            selected_found = True

            confirm = messagebox.askyesno("Confirm Deletion", "Are you sure you want to delete the selected row(s)?")
            if confirm:

                if user_account_student_table:
                    data_sheet = "Userdata"
                elif personal_student_table:
                    data_sheet = f"{personal_section_option.get()} Enrollment Information"
                elif family_student_table:
                    data_sheet = f"{family_section_option.get()} Enrollment Information"
                elif educ_student_table:
                    data_sheet = f"{educ_section_option.get()} Enrollment Information"

                try:
                    file_path_to_excel = "user_account_data.xlsx"
                    wb = load_workbook(file_path_to_excel)
                    if data_sheet not in wb.sheetnames:
                        messagebox.showerror("Error", f"Sheet '{data_sheet}' not found in the workbook.")
                        return

                    ws = wb[data_sheet]

                    for item in selected_item:
                        values = table.item(item, "values")
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                            if tuple(row) == values:
                                ws.delete_rows(row[0])  
                                break

                        table.delete(item)


                    wb.save(file_path_to_excel)
                    wb.close()
                    messagebox.showinfo("Deleted", "Selected row(s) have been deleted.")

                except FileNotFoundError:
                    messagebox.showerror("Error", "File not found.")
                except Exception as e:
                    messagebox.showerror("Error", f"An error occurred: {str(e)}")
   
            else:
                messagebox.showinfo("Canceled", "Deletion has been canceled.")
            break
    
    if not selected_found:
        messagebox.showwarning("No Selection", "Please select a row to delete.")

# Clear Button
def clear_all_rows():
    confirm = messagebox.askyesno("Confirm Clear All", "Are you sure you want to clear all rows?")
    if confirm:
        tables = [
            user_account_student_table,
            personal_student_table,
            family_student_table,
            educ_student_table
        ]

        try:
            file_path_to_excel = "user_account_data.xlsx"
            wb = load_workbook(file_path_to_excel)

            for table in tables:

                if user_account_student_table:
                    data_sheet = "Userdata"
                elif personal_student_table:
                    data_sheet = f"{personal_section_option.get()} Enrollment Information"
                elif family_student_table:
                    data_sheet = f"{family_section_option.get()} Enrollment Information"
                elif educ_student_table:
                    data_sheet = f"{educ_section_option.get()} Enrollment Information"
                else:
                    continue

                if data_sheet not in wb.sheetnames:
                    messagebox.showerror("Error", f"Sheet '{data_sheet}' not found in the workbook.")
                    return

                ws = wb[data_sheet]
                ws.delete_rows(2, ws.max_row)

                table.delete(*table.get_children())
                
            wb.save(file_path_to_excel)
            wb.close()
            messagebox.showinfo("Cleared", "All rows have been cleared.")    

        except FileNotFoundError:
            messagebox.showerror("Error", "File not found.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    else:
        messagebox.showinfo("Canceled", "Clear All has been canceled.")

# Admin Logo
def make_rounded_image(image_path, size, corner_radius):
    
    image = Image.open(image_path).convert("RGBA")
    image = image.resize(size, Image.Resampling.LANCZOS)
    
    mask = Image.new("L", size, 0)
    draw = ImageDraw.Draw(mask)
    draw.rounded_rectangle((0, 0, size[0], size[1]), radius=corner_radius, fill=255)

    rounded_image = ImageOps.fit(image, size, centering=(0.5, 0.5))
    rounded_image.putalpha(mask)

    return rounded_image


# ==================== Events ====================
def change_light_dark_mode_event(new_appearance_mode: str):
    ctk.set_appearance_mode(new_appearance_mode)

def show_frame(frame):
    frame.tkraise()
    if frame == personal_details:
        personal_details_show_data()
    elif frame == family_background:
        family_details_show_data()
    elif frame == educational_background:
        educational_details_show_data()

# ==================== Window Switch Function ====================
def SWITCH_WINDOW():
    window.destroy()
    subprocess.Popen(["python", "1_Portal_CTK.py"])

# ==================== Footer ====================
# ==================== Footer ====================
present = datetime.now().strftime("%B %d, %Y - %I:%M %p")

footer = CTkLabel(window, text=f"Admin Panel | Logged in as: Admin | © {present}", font=("Arial", 12))
footer.pack(side="bottom", fill="x", pady=5)

# ==================== Top Frame ====================
nav_top_frame = CTkFrame(window, fg_color="transparent")
nav_top_frame.pack(side="top", fill="x")

image_path = r".\assets\UniPass Admin.png"
rounded_img = make_rounded_image(image_path, size=(30, 30), corner_radius=30)
ctk_image = CTkImage(light_image=rounded_img, dark_image=rounded_img, size=(30, 30))
label = CTkLabel(nav_top_frame, image=ctk_image, text="")
label.pack(side="left", padx=20, pady=10)

logout_btn = CTkButton(nav_top_frame, text="Logout", font=("Arial", 15, "bold"), command=SWITCH_WINDOW)
logout_btn.pack(side="right", padx=5, pady=5)

switch_light_button = CTkOptionMenu(nav_top_frame, values=["Dark", "Light"] , command= change_light_dark_mode_event)
switch_light_button.pack(side="right", padx=5, pady=5)

# ==================== Main Container ====================
main_container = CTkFrame(window)
main_container.pack(fill="both", expand=True)

# ==================== Main Top Container ====================
main_top_container = CTkFrame(main_container, fg_color="transparent", bg_color="transparent")
main_top_container.pack(side="top", fill="x", padx=20, pady=20)

main_top_label = CTkLabel(main_top_container, text="Admin Panel", font=("Arial", 30, "bold"))
main_top_label.pack(pady=10, side="top")

# ==================== Main User Data Container ====================
user_data_container = CTkFrame(main_container)
user_data_container.pack(fill="both", expand=True)

# Right Frame: Manage Students
left_frame = CTkFrame(user_data_container, width=300, corner_radius=10)
left_frame.pack(side="left", fill="both", padx=20, pady=10)

left_title = CTkLabel(left_frame, text="Manage Students", font=("Arial", 25, "bold"))
left_title.pack(pady=10, padx=50)

left_user_account_label = CTkLabel(left_frame, text="User Account", font=("Arial", 18))
left_user_account_label.pack(anchor="w", padx=30, pady=(10, 0))
left_user_account_btn = CTkButton(left_frame, text="View Users Account", font=("Arial", 14), width=200, command=lambda: show_frame(user_account))
left_user_account_btn.pack(anchor="w", padx=30, pady=(5, 10))

left_personal_label = CTkLabel(left_frame, text="Personal Background", font=("Arial", 18))
left_personal_label.pack(anchor="w", padx=30, pady=(10, 0))
left_personal_btn = CTkButton(left_frame, text="View Personal Background", font=("Arial", 14), width=200, command=lambda: show_frame(personal_details))
left_personal_btn.pack(anchor="w", padx=30, pady=(5, 10))

left_family_label = CTkLabel(left_frame, text="Family Background", font=("Arial", 18))
left_family_label.pack(anchor="w", padx=30, pady=(10, 0))
left_family_btn = CTkButton(left_frame, text="View Family Background", font=("Arial", 14), width=200, command=lambda: show_frame(family_background))
left_family_btn.pack(anchor="w", padx=30, pady=(5, 10))

left_educational_label = CTkLabel(left_frame, text="Educational Background", font=("Arial", 18))
left_educational_label.pack(anchor="w", padx=30, pady=(10, 0))
left_educational_btn = CTkButton(left_frame, text="View Educational Background", font=("Arial", 14), width=200, command=lambda: show_frame(educational_background))
left_educational_btn.pack(anchor="w", padx=30, pady=(5, 10))

left_attendance_label = CTkLabel(left_frame, text="Attendance Record", font=("Arial", 18))
left_attendance_label.pack(anchor="w", padx=30, pady=(10, 0))
left_attendance_btn = CTkButton(left_frame, text="View Attendance Records", font=("Arial", 14), width=200, command=lambda: show_frame(attendance))
left_attendance_btn.pack(anchor="w", padx=30, pady=(5, 10))

pick_file_btn = CTkButton(left_frame, text="Import Data", font=("Arial", 14), width=200, command=user_account_show_data)
pick_file_btn.pack(side="bottom", padx=30, pady=(5, 10))

right_frame = CTkFrame(user_data_container, width=1000, corner_radius=10, fg_color="transparent")
right_frame.pack(side="right", fill="both", expand=True, padx=20, pady=10)

user_account = CTkFrame(right_frame, corner_radius=10)
personal_details = CTkFrame(right_frame, corner_radius=10)
family_background = CTkFrame(right_frame, corner_radius=10)
educational_background = CTkFrame(right_frame, corner_radius=10)
attendance = CTkFrame(right_frame, corner_radius=10)

for frame in (user_account, personal_details, family_background, educational_background, attendance):
    frame.place(relx=0, rely=0, relwidth=1, relheight=1)

show_frame(user_account)

btn_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
btn_frame.pack(pady=10, side="bottom")
buttons = ["Delete", "Clear"]
for index, btn in enumerate(buttons):
    row = index // 2
    column = index % 2

    if btn == "Delete":
        command = delete_selected_row
    else:
        command = clear_all_rows
    ctk.CTkButton(btn_frame, text=btn, width=150, command=command).grid(row=row, column=column, padx=10, pady=5)




# ==================== USER ACCOUNT FRAME ====================
user_account_right_title = CTkLabel(user_account, text="Account Data", font=("Arial", 25, "bold"))
user_account_right_title.pack(pady=10, padx=50)

user_account_right_top_frame = CTkFrame(user_account, fg_color="transparent")
user_account_right_top_frame.pack(side="top", fill="x", padx=20, pady=10)

user_account_search_option = CTkOptionMenu(user_account_right_top_frame, values=["Username", "Email"], width=200, height=30)
user_account_search_option.set("Search By")
user_account_search_option.grid(row=0, column=0, padx=5, pady=5)

user_account_search_entry = CTkEntry(user_account_right_top_frame, width=400, height=30, placeholder_text="Search Here", font=("Arial", 14))
user_account_search_entry.grid(row=0, column=1, padx=5, pady=5)

user_account_search_btn = CTkButton(user_account_right_top_frame, text="Search", width=100, height=30, command=user_account_search)
user_account_search_btn.grid(row=0, column=2, padx=5, pady=5)

user_account_show_all_btn = CTkButton(user_account_right_top_frame, text="Show All", width=100, height=30, command=show_all_user_account_data)
user_account_show_all_btn.grid(row=0, column=3, padx=5, pady=5)

# User Account Table
user_account_table_frame = CTkFrame(user_account)
user_account_table_frame.pack(side="top", fill="both", expand=True, padx=20, pady=10)

cols = (
        "First Name", "Last Name", "Email", "Username", "Password"
        )
user_account_student_table = ttk.Treeview(user_account_table_frame, columns=cols, show="headings")

for col in cols:
    user_account_student_table.heading(col, text=col.capitalize())
    user_account_student_table.column(col, width=120)

vertical_scrollbar = ttk.Scrollbar(user_account_table_frame, orient="vertical", command=user_account_student_table.yview)
vertical_scrollbar.pack(side="right", fill="y")
user_account_student_table.configure(yscrollcommand=vertical_scrollbar.set)

horizontal_scrollbar = ttk.Scrollbar(user_account_table_frame, orient="horizontal", command=user_account_student_table.xview)
horizontal_scrollbar.pack(side="bottom", fill="x")
user_account_student_table.configure(xscrollcommand=horizontal_scrollbar.set)

user_account_student_table.pack(fill="both", expand=True)


# ==================== PERSONAL BACKGROUND FRAME ====================
personal_right_title = CTkLabel(personal_details, text="Student Data", font=("Arial", 25, "bold"))
personal_right_title.pack(pady=10, padx=50)

personal_right_top_frame = CTkFrame(personal_details, fg_color="transparent")
personal_right_top_frame.pack(side="top", fill="x", padx=20, pady=10)

personal_section_option = CTkOptionMenu(personal_right_top_frame, values=["1A", "1B", "1C"], width=125, height=30, command=lambda _: personal_details_show_data())
personal_section_option.set("1A")
personal_section_option.grid(row=0, column=0, padx=5, pady=5)

personal_search_option = CTkOptionMenu(personal_right_top_frame, values=["Student ID", "Surname"], width=125, height=30)
personal_search_option.set("Search By")
personal_search_option.grid(row=0, column=1, padx=5, pady=5)

personal_search_entry = CTkEntry(personal_right_top_frame, width=340, height=30, placeholder_text="Search Here", font=("Arial", 14))
personal_search_entry.grid(row=0, column=2, padx=5, pady=5)

personal_search_btn = CTkButton(personal_right_top_frame, text="Search", width=100, height=30,command=personal_details_search)
personal_search_btn.grid(row=0, column=3, padx=5, pady=5)

personal_show_all_btn = CTkButton(personal_right_top_frame, text="Show All", width=100, height=30, command=personal_details_show_data)
personal_show_all_btn.grid(row=0, column=4, padx=5, pady=5)

# Personal Account Table
personal_table_frame = CTkFrame(personal_details)
personal_table_frame.pack(side="top", fill="both", expand=True, padx=20, pady=10)

cols = (
        "Student ID", "Course/Section", "LRN", 
        "",
        "Surname", "Firstname", "Middle Initial", 
        "",
        "Gender", "Age", "Birthdate", "Birthplace", "Nationality", "Religion", "Marital Status", "Language Spoken",
        "",
        "Street", "Barangay", "City", "Zip Code", "Province", "Country", 
        "",
        "Email", "Contact Number"
        )

personal_student_table = ttk.Treeview(personal_table_frame, columns=cols, show="headings")

for col in cols:
    personal_student_table.heading(col, text=col.capitalize())
    personal_student_table.column(col, width=120)

vertical_scrollbar = ttk.Scrollbar(personal_table_frame, orient="vertical", command=personal_student_table.yview)
vertical_scrollbar.pack(side="right", fill="y")
personal_student_table.configure(yscrollcommand=vertical_scrollbar.set)

horizontal_scrollbar = ttk.Scrollbar(personal_table_frame, orient="horizontal", command=personal_student_table.xview)
horizontal_scrollbar.pack(side="bottom", fill="x")
personal_student_table.configure(xscrollcommand=horizontal_scrollbar.set)

personal_student_table.pack(fill="both", expand=True)


# ==================== FAMILY BACKGROUND FRAME ====================
family_right_title = CTkLabel(family_background, text="Family Data", font=("Arial", 25, "bold"))
family_right_title.pack(pady=10, padx=50)

family_right_top_frame = CTkFrame(family_background, fg_color="transparent")
family_right_top_frame.pack(side="top", fill="x", padx=20, pady=10)

family_section_option = CTkOptionMenu(family_right_top_frame, values=["1A", "1B", "1C"], width=125, height=30, command=lambda _: family_details_show_data())
family_section_option.set("1A")
family_section_option.grid(row=0, column=0, padx=5, pady=5)

family_search_option = CTkOptionMenu(family_right_top_frame, values=["Father's Name", "Mother's Name", "Guardian's Name"], width=125, height=30)
family_search_option.set("Search By")
family_search_option.grid(row=0, column=1, padx=5, pady=5)

family_search_entry = CTkEntry(family_right_top_frame, width=340, height=30, placeholder_text="Search Here", font=("Arial", 14))
family_search_entry.grid(row=0, column=2, padx=5, pady=5)

family_search_btn = CTkButton(family_right_top_frame, text="Search", width=100, height=30, command=family_details_search)   #NOT FUNCTIONING PROPERLY
family_search_btn.grid(row=0, column=3, padx=5, pady=5)

family_show_all_btn = CTkButton(family_right_top_frame, text="Show All", width=100, height=30, command=family_details_show_data)
family_show_all_btn.grid(row=0, column=4, padx=5, pady=5)

# User Account Table
family_table_frame = CTkFrame(family_background)
family_table_frame.pack(side="top", fill="both", expand=True, padx=20, pady=10)

cols = (                
        "Father's Name", "Father's Address", "Father's Occupation", "Father's Contact No",
        "",
        "Mother's Name", "Mother's Address", "Mother's Occupation", "Mother's Contact No",
        "",
        "Guardian's Name", "Guardian's Relationship", "Guardian's Address", "Guardian's Occupation", "Guardian's Contact No",
        )

family_student_table = ttk.Treeview(family_table_frame, columns=cols, show="headings")

for col in cols:
    family_student_table.heading(col, text=col.capitalize())
    family_student_table.column(col, width=150)

vertical_scrollbar = ttk.Scrollbar(family_table_frame, orient="vertical", command=family_student_table.yview)
vertical_scrollbar.pack(side="right", fill="y")
family_student_table.configure(yscrollcommand=vertical_scrollbar.set)

horizontal_scrollbar = ttk.Scrollbar(family_table_frame, orient="horizontal", command=family_student_table.xview)
horizontal_scrollbar.pack(side="bottom", fill="x")
family_student_table.configure(xscrollcommand=horizontal_scrollbar.set)

family_student_table.pack(fill="both", expand=True)


# ==================== FAMILY BACKGROUND FRAME ====================
educ_right_title = CTkLabel(educational_background, text="Educational Data", font=("Arial", 25, "bold"))
educ_right_title.pack(pady=10, padx=50)

educ_right_top_frame = CTkFrame(educational_background, fg_color="transparent")
educ_right_top_frame.pack(side="top", fill="x", padx=20, pady=10)

educ_section_option = CTkOptionMenu(educ_right_top_frame, values=["1A", "1B", "1C"], width=125, height=30, command=lambda _: educational_details_show_data())
educ_section_option.set("1A")
educ_section_option.grid(row=0, column=0, padx=5, pady=5)

educ_search_option = CTkOptionMenu(educ_right_top_frame, values=["ELEM", "JHS", "SHS", "COL"], width=125, height=30)
educ_search_option.set("Search By")
educ_search_option.grid(row=0, column=1, padx=5, pady=5)

educ_search_entry = CTkEntry(educ_right_top_frame, width=340, height=30, placeholder_text="Search Here", font=("Arial", 14))
educ_search_entry.grid(row=0, column=2, padx=5, pady=5)

educ_search_btn = CTkButton(educ_right_top_frame, text="Search", width=100, height=30, command=educational_details_search)
educ_search_btn.grid(row=0, column=3, padx=5, pady=5)

educ_show_all_btn = CTkButton(educ_right_top_frame, text="Show All", width=100, height=30, command=educational_details_show_data)
educ_show_all_btn.grid(row=0, column=4, padx=5, pady=5)

# User Account Table
educ_table_frame = CTkFrame(educational_background)
educ_table_frame.pack(side="top", fill="both", expand=True, padx=20, pady=10)

cols = (                
        "Elementary School", "Elementary Address", "Elementary Year Graduated",
        "",
        "Junior High School", "Junior High Address", "Junior High Year Graduated",
        "",
        "Senior High School", "Senior High Address", "Senior High Strand", "Senior High Year Graduated",
        "",
        "College", "College Address", "College Year Graduated"
        )

educ_student_table = ttk.Treeview(educ_table_frame, columns=cols, show="headings")

for col in cols:
    educ_student_table.heading(col, text=col.capitalize())
    educ_student_table.column(col, width=150)

vertical_scrollbar = ttk.Scrollbar(educ_table_frame, orient="vertical", command=educ_student_table.yview)
vertical_scrollbar.pack(side="right", fill="y")
educ_student_table.configure(yscrollcommand=vertical_scrollbar.set)

horizontal_scrollbar = ttk.Scrollbar(educ_table_frame, orient="horizontal", command=educ_student_table.xview)
horizontal_scrollbar.pack(side="bottom", fill="x")
educ_student_table.configure(xscrollcommand=horizontal_scrollbar.set)
educ_student_table.pack(fill="both", expand=True)

style = ttk.Style()
style.configure("Treeview", background="#2b2b2b", foreground="white", fieldbackground="#2b2b2b")


# ==================== Window Starter ====================
window.columnconfigure(0, weight=1)
window.rowconfigure(0, weight=1)
window.mainloop()